# This script was created by Jordan Nielson to faciliate faster parsing of the creative evaluations for CS260
# Originally created with openpyxl 2.3.5 and python 2.7

from __future__ import division
import re
import os
import openpyxl
# Averages workbook (to make it easy to copy/paste the eval averages)
averages = openpyxl.Workbook()
averageSheet = averages.get_sheet_by_name('Sheet')

# input the path to the evaluations file
path = raw_input("Please enter the file name: ")
# It doesn't have to end in .xlsx when input... since sometimes I'm lazy.
regex = r'xlsx$'
if not re.search(regex, path):
    path += '.xlsx'
# Either way by this point it should end in .xlsx
print(path)
# Load the spreadsheet into memory with openpyxl
wb = openpyxl.load_workbook(path)
# Create a new workbook for the list of evaluators
result = openpyxl.Workbook()
# Get the sheet with the form responses
sheet = wb.get_sheet_by_name('Form Responses 1')
# print sheet.title
# Why is lastrow sheet.max_row + 1? Ranges dont include the end value
lastrow = sheet.max_row + 1
# Keep track of how many groups there are
maxGroupNum = 0
# evaluators kept in this set
evaluatorNames = set()
# Group Names are kept here in order to place into the averages sheet
groupNames = set()
for i in range(2, lastrow):
    name = sheet.cell(row=i, column=2).value
    evaluatorNames.add(name)
    groupName = sheet.cell(row=i, column=3).value
    Group = groupName.split(')')
    groupNames.add(Group[1])
    #print(sheet.cell(row=i, column=3).value)
    if int(Group[0]) > maxGroupNum:
        maxGroupNum = int(Group[0])
    # Replace the full group with just the number
    sheet.cell(row=i, column=3).value = Group[0]
    # Replace the evaluator name with the group names to keep feedback
    # anonymous
    sheet.cell(row=i, column=2).value = Group[1]
# Sort and save the evaluatorNames
evaluatorNames = sorted(evaluatorNames)
i = 1
for x in evaluatorNames:
    # place each on a row
    result.active.cell(row=i, column=1).value = x
    i += 1
# save the list of evaluators to a file
result.save('evaluators' + path)

# Keep track of which row each sheet is on
groupSheetRow = {}


numColumnsToCopy = sheet.max_column + 1

# Create sheets for each group with the header row
for i in range(1, maxGroupNum + 1):
    wb.create_sheet(title=str(i))
    groupSheetRow[str(i)] = 2
    newSheet = wb.get_sheet_by_name(str(i))
    for y in range(2, numColumnsToCopy):
        newSheet.cell(row=1, column=y).value = sheet.cell(
            row=1, column=y).value

# Grab the first group to start the currentGroup loop
currentGroup = sheet.cell(row=2, column=3).value


# Grab the sheet for the current group
groupSheet = wb.get_sheet_by_name(currentGroup)

for i in range(2, lastrow):
    # Copy the current row to the group's sheet
    for y in range(2, numColumnsToCopy):
        groupSheet.cell(row=groupSheetRow[currentGroup], column=y).value = sheet.cell(
            row=i, column=y).value
    # Keep track of which row the group's sheet is on next
    groupSheetRow[currentGroup] += 1
    # Check to see if you need to change sheets
    nextGroup = sheet.cell(row=i + 1, column=3).value
    if nextGroup != currentGroup:
        if nextGroup is not None:
             # print "Regardless... we should hit here."
            # Copy the current row to the group's sheet
            currentGroup = nextGroup
            groupSheet = wb.get_sheet_by_name(currentGroup)
        else:
            print("You've either reached the end or have an error!")

# Grab the list of sheets in the workbook
groupSheets = wb.get_sheet_names()
# Happens to include the original sheet, so the length is what we need to go to
for group in range(1, len(groupSheets)):
    # Grab the group's sheet
    groupSheet = wb.get_sheet_by_name(groupSheets[group])
    # Grab the current row of the group's sheet
    currentRow = groupSheetRow[groupSheets[group]]
    if currentRow == 2:
        print("There's nothing in this sheet! (aka the group had no evaluations)")
    else:
        # Create the Average row (currently hard-coded for the form as of
        # 9/4/17):
        averageValues = 0
        for y in range(4, 5): #TODO: UPDATE IF FORM CHANGES
            Letter = openpyxl.cell.get_column_letter(y)
            groupSheet.cell(row=currentRow, column=y).value = '=AVERAGE(' + \
                Letter + '2:' + Letter + str(currentRow - 1) + ')'

            # Calculate the average value for the copy/paste sheet
            sum = 0
            numRows = 0
            for z in range(2, currentRow):
                if groupSheet.cell(row=z, column=y).value is None:
                    continue
                sum += groupSheet.cell(row=z, column=y).value
                numRows += 1
            if currentRow == 3:
                if groupSheet.cell(row=2, column=y).value is None:
                    continue
                sum += groupSheet.cell(row=2, column=y).value
                numRows += 1
            average = sum / numRows
            averageValues += average
            averageSheet.cell(row=group, column=y).value = sum / numRows

        groupSheet.cell(row=currentRow, column=2).value = "AVERAGES:"
        # groupSheet.cell(row=currentRow, column=3).value = '=AVERAGE(D' + \
            # str(currentRow) + ':J' + str(currentRow) + ')'
        averageSheet.cell(row=group, column=2).value = "AVERAGES:"
        # averageSheet.cell(row=group, column=3).value = averageValues / 7

wb.save("Parsed" + path)

dir_name = raw_input(
    "Please enter the name of the folder you want the groups in: ")
try:
    os.mkdir(dir_name)
except OSError as e:
    print(e)


# Create a file for each group
for group in range(1, len(groupSheets)):
    newWB = openpyxl.Workbook()
    newSheet = newWB.get_sheet_by_name('Sheet')
    groupSheet = wb.get_sheet_by_name(groupSheets[group])
    groupName = groupSheet.cell(row=2, column=2).value
    # Check to make sure it's actually a group
    if groupName is None:
        continue
    # Copy their sheet to their file
    averageSheet.cell(row=group, column=1).value = groupSheets[
        group] + ') ' + groupName
    for x in range(1, groupSheet.max_row + 1):
        for y in range(1, groupSheet.max_column + 1):
            newSheet.cell(row=x, column=y).value = groupSheet.cell(
                row=x, column=y).value
    newWB.save(dir_name + '/' + groupSheets[group] + ' ' + groupName + '.xlsx')

# Add functionality to copy all averages to one sheet for easy copy/paste
# I'm sorry that this doesn't output them in any coherent order...
averages.save('averages' + path)
