# This script was created by Jordan Nielson to faciliate faster parsing of the class activities for CS260
# Originally created with openpyxl 2.3.5 and python 2.7

import re
import openpyxl

# input the PATH to the evaluations file
PATH = raw_input("Please enter the file name: ")
# It doesn't have to end in .xlsx when input... since sometimes I'm lazy.
REGEX = r'xlsx$'
if not re.search(REGEX, PATH):
    PATH += '.xlsx'
# Either way by this point it should end in .xlsx
print PATH
# Load the spreadsheet into memory with openpyxl
WB = openpyxl.load_workbook(PATH)
# Create a new workbook for the list of evaluators
RESULT = openpyxl.Workbook()
# Get the sheet with the form responses
sheet = WB.get_sheet_by_name('Sheet1')
# print sheet.title
# Why is lastrow sheet.max_row + 1? Ranges dont include the end value
lastrow = sheet.max_row + 1
# The names of the people kept in this set
names = set()
for i in range(2, lastrow):
    name = sheet.cell(row=i, column=1).value
    names.add(name)
# Sort and save the names
names = sorted(names)
i = 1
for x in names:
    # place each on a row
    RESULT.active.cell(row=i, column=1).value = x
    i += 1
# save the list of evaluators to a file
RESULT.save('activity'+PATH)
